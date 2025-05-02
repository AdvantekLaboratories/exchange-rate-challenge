#!/usr/bin/env python3

import argparse
import csv
import datetime
import json
import os
import statistics
import sys
from typing import List, Dict, Tuple, Optional, Any

import requests
from bs4 import BeautifulSoup
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook


class ExchangeRateTracker:
    
    def __init__(self, csv_file: str = 'rates.csv'):
        self.csv_file = csv_file
        self.ensure_csv_exists()
        
        self.api_sources = {
            'ecb': self._fetch_from_ecb,
            'mnb': self._fetch_from_mnb
        }
    
    def ensure_csv_exists(self) -> None:
        if not os.path.exists(self.csv_file):
            with open(self.csv_file, 'w', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(['date', 'rate', 'source'])
    
    def fetch_current_rate(self, source: str = 'auto') -> Tuple[datetime.date, float, str]:
        if source == 'auto':
            sources_to_try = ['ecb', 'mnb']
            
            for s in sources_to_try:
                try:
                    date, rate, source_used = self.api_sources[s]()
                    print(f"[✔] Successfully fetched rate from {source_used}")
                    return date, rate, source_used
                except Exception as e:
                    print(f"[⚠️] Failed to fetch from {s}: {e}")
                    continue
            
            print("[❌] Failed to fetch exchange rate from any source")
            sys.exit(1)
        else:
            if source not in self.api_sources:
                print(f"[❌] Unknown source: {source}")
                print(f"[ℹ️] Available sources: {', '.join(self.api_sources.keys())}")
                sys.exit(1)
            
            try:
                return self.api_sources[source]()
            except Exception as e:
                print(f"[❌] Failed to fetch from {source}: {e}")
                sys.exit(1)
    
    def _fetch_from_ecb(self) -> Tuple[datetime.date, float, str]:
        url = "https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'xml')
        huf_rate_elem = soup.find('Cube', currency='HUF')
        
        if not huf_rate_elem:
            raise ValueError("Could not find HUF rate in the XML data")
        
        rate = float(huf_rate_elem['rate'])
        today = datetime.date.today()
        
        return today, rate, 'ecb'
    
    def _fetch_from_mnb(self) -> Tuple[datetime.date, float, str]:
        url = "https://www.mnb.hu/en/arfolyamok"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        rate_elem = soup.select_one("table.exchange-rates tr:contains('EUR') td:nth-child(3)")
        
        if not rate_elem:
            raise ValueError("Could not find EUR/HUF rate on the MNB website")
        
        rate_text = rate_elem.text.strip().replace(',', '.')
        rate = float(rate_text)
        today = datetime.date.today()
        
        return today, rate, 'mnb'
    
    def store_rate(self, date: datetime.date, rate: float, source: str) -> None:
        with open(self.csv_file, 'a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([date.isoformat(), rate, source])
    
    def load_historical_rates(self) -> List[Dict[str, Any]]:
        rates = []
        try:
            with open(self.csv_file, 'r', newline='') as file:
                reader = csv.DictReader(file)
                rates = list(reader)
            
            for entry in rates:
                entry['rate'] = float(entry['rate'])
                if 'source' not in entry:
                    entry['source'] = 'unknown'
            
            return rates
        except FileNotFoundError:
            print(f"[❌] Error: Could not find {self.csv_file}")
            sys.exit(1)
    
    def calculate_moving_average(self, days: int = 7) -> Optional[float]:
        rates = self.load_historical_rates()
        
        if len(rates) < days:
            print(f"[⚠️] Warning: Not enough data for {days}-day moving average. Need at least {days} days of data.")
            return None
        
        recent_rates = [entry['rate'] for entry in rates[-days:]]
        return statistics.mean(recent_rates)
    
    def generate_recommendation(self) -> Tuple[float, float, str]:
        rates = self.load_historical_rates()
        
        if not rates:
            print("[❌] Error: No historical data available. Run 'fetch' command first.")
            sys.exit(1)
        
        ma_7day = self.calculate_moving_average(7)
        
        if ma_7day is None:
            print("[⚠️] Warning: Using available data for recommendation.")
            ma_7day = statistics.mean([entry['rate'] for entry in rates])
        
        current_rate = rates[-1]['rate']
        
        if current_rate > ma_7day:
            recommendation = "BUY EUR"
        else:
            recommendation = "SELL EUR"
        
        return ma_7day, current_rate, recommendation
    
    def plot_historical_rates(self, output_file: str = 'exchange_rates.png') -> None:
        rates = self.load_historical_rates()
        
        if len(rates) < 2:
            print("[❌] Error: Not enough data to create a plot.")
            return
        
        df = pd.DataFrame(rates)
        df['date'] = pd.to_datetime(df['date'])
        df.set_index('date', inplace=True)
        
        if len(df) >= 7:
            df['ma_7day'] = df['rate'].rolling(window=7).mean()
            
            plt.figure(figsize=(10, 6))
            plt.plot(df.index, df['rate'], label='EUR/HUF Rate', marker='o')
            plt.plot(df.index, df['ma_7day'], label='7-day Moving Average', linestyle='--')
            plt.title('EUR/HUF Exchange Rate History')
            plt.xlabel('Date')
            plt.ylabel('Exchange Rate (HUF/EUR)')
            plt.legend()
            plt.grid(True)
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            plt.savefig(output_file)
            print(f"[✔] Plot saved as {output_file}")
        else:
            plt.figure(figsize=(10, 6))
            plt.plot(df.index, df['rate'], label='EUR/HUF Rate', marker='o')
            plt.title('EUR/HUF Exchange Rate History')
            plt.xlabel('Date')
            plt.ylabel('Exchange Rate (HUF/EUR)')
            plt.grid(True)
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            plt.savefig(output_file)
            print(f"[✔] Plot saved as {output_file}")
    
    def export_data(self, format_type: str, output_file: Optional[str] = None) -> None:
        rates = self.load_historical_rates()
        
        if not rates:
            print("[❌] Error: No data available to export.")
            return
        
        df = pd.DataFrame(rates)
        df['date'] = pd.to_datetime(df['date'])
        
        if output_file is None:
            today = datetime.date.today().isoformat()
            output_file = f"exchange_rates_{today}.{format_type}"
            if format_type == 'excel':
                output_file = f"exchange_rates_{today}.xlsx"
        
        if format_type.lower() == 'csv':
            df.to_csv(output_file, index=False)
            print(f"[✔] Data exported to CSV: {output_file}")
        
        elif format_type.lower() == 'json':
            df['date'] = df['date'].dt.strftime('%Y-%m-%d')
            with open(output_file, 'w') as f:
                json.dump(df.to_dict(orient='records'), f, indent=2)
            print(f"[✔] Data exported to JSON: {output_file}")
        
        elif format_type.lower() == 'excel':
            df.to_excel(output_file, index=False, sheet_name='Exchange Rates')
            print(f"[✔] Data exported to Excel: {output_file}")
        
        else:
            print(f"[❌] Unsupported export format: {format_type}")
            print("[ℹ️] Supported formats: csv, json, excel")
    
    def compare_historical_data(self, period1: Optional[str] = None, period2: Optional[str] = None,
                               days1: int = 7, days2: int = 7, export_format: Optional[str] = None) -> dict:
        rates = self.load_historical_rates()
        
        if len(rates) < max(days1, days2):
            print(f"[⚠️] Warning: Not enough data for comparison. Need at least {max(days1, days2)} days of data.")
            return {}
        
        df = pd.DataFrame(rates)
        df['date'] = pd.to_datetime(df['date'])
        df.sort_values('date', inplace=True)
        
        if period1 is None:
            period1_data = df.iloc[-days1:]
        else:
            try:
                start_date = pd.to_datetime(period1)
                period1_data = df[df['date'] >= start_date].head(days1)
            except ValueError:
                print(f"[❌] Invalid date format for period1: {period1}. Use YYYY-MM-DD.")
                return {}
        
        if period2 is None:
            if period1 is None:
                period2_data = df.iloc[-(days1+days2):-days1]
            else:
                period2_end = period1_data.iloc[0]['date'] - pd.Timedelta(days=1)
                period2_data = df[df['date'] <= period2_end].tail(days2)
        else:
            try:
                start_date = pd.to_datetime(period2)
                period2_data = df[df['date'] >= start_date].head(days2)
            except ValueError:
                print(f"[❌] Invalid date format for period2: {period2}. Use YYYY-MM-DD.")
                return {}
        
        period1_stats = {
            'start_date': period1_data['date'].min().strftime('%Y-%m-%d'),
            'end_date': period1_data['date'].max().strftime('%Y-%m-%d'),
            'days': len(period1_data),
            'avg_rate': round(period1_data['rate'].mean(), 2),
            'min_rate': round(period1_data['rate'].min(), 2),
            'max_rate': round(period1_data['rate'].max(), 2),
            'volatility': round(period1_data['rate'].std(), 2)
        }
        
        period2_stats = {
            'start_date': period2_data['date'].min().strftime('%Y-%m-%d'),
            'end_date': period2_data['date'].max().strftime('%Y-%m-%d'),
            'days': len(period2_data),
            'avg_rate': round(period2_data['rate'].mean(), 2),
            'min_rate': round(period2_data['rate'].min(), 2),
            'max_rate': round(period2_data['rate'].max(), 2),
            'volatility': round(period2_data['rate'].std(), 2)
        }
        
        comparison = {
            'avg_rate_diff': round(period1_stats['avg_rate'] - period2_stats['avg_rate'], 2),
            'avg_rate_pct_change': round((period1_stats['avg_rate'] / period2_stats['avg_rate'] - 1) * 100, 2),
            'volatility_change': round(period1_stats['volatility'] - period2_stats['volatility'], 2),
            'volatility_pct_change': round((period1_stats['volatility'] / period2_stats['volatility'] - 1) * 100, 2) if period2_stats['volatility'] > 0 else float('inf')
        }
        
        results = {
            'period1': period1_stats,
            'period2': period2_stats,
            'comparison': comparison
        }
        
        if export_format:
            self._export_comparison(results, export_format)
        
        return results
    
    def _export_comparison(self, comparison_data: dict, format_type: str) -> None:
        today = datetime.date.today().isoformat()
        output_file = f"exchange_rate_comparison_{today}.{format_type}"
        
        if format_type == 'excel':
            output_file = f"exchange_rate_comparison_{today}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Comparison"
            
            ws['A1'] = "Metric"
            ws['B1'] = f"Period 1 ({comparison_data['period1']['start_date']} to {comparison_data['period1']['end_date']})"
            ws['C1'] = f"Period 2 ({comparison_data['period2']['start_date']} to {comparison_data['period2']['end_date']})"
            ws['D1'] = "Difference"
            ws['E1'] = "% Change"
            
            metrics = [
                ('Average Rate', 'avg_rate'),
                ('Minimum Rate', 'min_rate'),
                ('Maximum Rate', 'max_rate'),
                ('Volatility', 'volatility'),
                ('Days', 'days')
            ]
            
            for idx, (label, key) in enumerate(metrics, 2):
                ws[f'A{idx}'] = label
                ws[f'B{idx}'] = comparison_data['period1'][key]
                ws[f'C{idx}'] = comparison_data['period2'][key]
                
                if key in ['avg_rate', 'min_rate', 'max_rate', 'volatility']:
                    diff = comparison_data['period1'][key] - comparison_data['period2'][key]
                    pct_change = (comparison_data['period1'][key] / comparison_data['period2'][key] - 1) * 100 if comparison_data['period2'][key] != 0 else float('inf')
                    
                    ws[f'D{idx}'] = round(diff, 2)
                    ws[f'E{idx}'] = f"{round(pct_change, 2)}%"
            
            wb.save(output_file)
            print(f"[✔] Comparison exported to Excel: {output_file}")
            
        elif format_type == 'json':
            with open(output_file, 'w') as f:
                json.dump(comparison_data, f, indent=2)
            print(f"[✔] Comparison exported to JSON: {output_file}")
            
        elif format_type == 'csv':
            rows = []
            headers = ['metric', 'period1', 'period2', 'difference', 'pct_change']
            
            metrics = [
                ('Average Rate', 'avg_rate'),
                ('Minimum Rate', 'min_rate'),
                ('Maximum Rate', 'max_rate'),
                ('Volatility', 'volatility'),
                ('Days', 'days')
            ]
            
            for label, key in metrics:
                diff = comparison_data['period1'][key] - comparison_data['period2'][key] if key != 'days' else 'N/A'
                if key != 'days' and comparison_data['period2'][key] != 0:
                    pct_change = f"{round((comparison_data['period1'][key] / comparison_data['period2'][key] - 1) * 100, 2)}%"
                else:
                    pct_change = 'N/A'
                
                rows.append({
                    'metric': label,
                    'period1': comparison_data['period1'][key],
                    'period2': comparison_data['period2'][key],
                    'difference': diff,
                    'pct_change': pct_change
                })
            
            rows.insert(0, {
                'metric': 'Date Range',
                'period1': f"{comparison_data['period1']['start_date']} to {comparison_data['period1']['end_date']}",
                'period2': f"{comparison_data['period2']['start_date']} to {comparison_data['period2']['end_date']}",
                'difference': 'N/A',
                'pct_change': 'N/A'
            })
            
            with open(output_file, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(rows)
            
            print(f"[✔] Comparison exported to CSV: {output_file}")
            
        else:
            print(f"[❌] Unsupported export format: {format_type}")
            print("[ℹ️] Supported formats: csv, json, excel")


def setup_argparse() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description='EUR/HUF Exchange Rate Tracker and Recommendation Tool',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    subparsers = parser.add_subparsers(dest='command', help='Command to run')
    
    fetch_parser = subparsers.add_parser('fetch', help='Fetch and store today\'s exchange rate')
    fetch_parser.add_argument(
        '--source', '-s',
        choices=['ecb', 'mnb', 'auto'],
        default='auto',
        help='Source to fetch the exchange rate from (default: auto)'
    )
    
    recommend_parser = subparsers.add_parser('recommend', help='Analyze rates and provide a trading recommendation')
    
    plot_parser = subparsers.add_parser('plot', help='Create a plot of historical exchange rates')
    plot_parser.add_argument(
        '--output', '-o',
        help='Output filename for the plot',
        default='exchange_rates.png'
    )
    
    export_parser = subparsers.add_parser('export', help='Export historical data to various formats')
    export_parser.add_argument(
        '--format', '-f',
        choices=['csv', 'json', 'excel'],
        required=True,
        help='Format to export to'
    )
    export_parser.add_argument(
        '--output', '-o',
        help='Output filename (default: exchange_rates_YYYY-MM-DD.ext)'
    )
    
    compare_parser = subparsers.add_parser('compare', help='Compare exchange rates between two periods')
    compare_parser.add_argument(
        '--period1', '-p1',
        help='Start date for first period (YYYY-MM-DD), defaults to last N days'
    )
    compare_parser.add_argument(
        '--period2', '-p2',
        help='Start date for second period (YYYY-MM-DD), defaults to N days before period1'
    )
    compare_parser.add_argument(
        '--days1', '-d1',
        type=int,
        default=7,
        help='Number of days for first period if period1 is not specified (default: 7)'
    )
    compare_parser.add_argument(
        '--days2', '-d2',
        type=int,
        default=7,
        help='Number of days for second period if period2 is not specified (default: 7)'
    )
    compare_parser.add_argument(
        '--export', '-e',
        choices=['csv', 'json', 'excel'],
        help='Export comparison to the specified format'
    )
    
    sources_parser = subparsers.add_parser('sources', help='List available data sources')
    
    return parser


def main():
    parser = setup_argparse()
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return
    
    tracker = ExchangeRateTracker()
    
    if args.command == 'fetch':
        date, rate, source = tracker.fetch_current_rate(args.source)
        tracker.store_rate(date, rate, source)
        print(f"[✔] Stored rate: {rate:.2f} HUF/EUR on {date.isoformat()} (source: {source})")
    
    elif args.command == 'recommend':
        ma_7day, current_rate, recommendation = tracker.generate_recommendation()
        print(f"📊 7-day Moving Average: {ma_7day:.2f}")
        print(f"📈 Today's Rate: {current_rate:.2f}")
        print(f"💡 Recommendation: {recommendation}")
    
    elif args.command == 'plot':
        tracker.plot_historical_rates(args.output)
    
    elif args.command == 'export':
        tracker.export_data(args.format, args.output)
    
    elif args.command == 'compare':
        results = tracker.compare_historical_data(
            period1=args.period1,
            period2=args.period2,
            days1=args.days1,
            days2=args.days2,
            export_format=args.export
        )
        
        if results:
            print("\n📊 Exchange Rate Comparison 📊")
            print(f"\nPeriod 1 ({results['period1']['start_date']} to {results['period1']['end_date']}, {results['period1']['days']} days):")
            print(f"  Average Rate: {results['period1']['avg_rate']:.2f} HUF/EUR")
            print(f"  Min/Max: {results['period1']['min_rate']:.2f} - {results['period1']['max_rate']:.2f} HUF/EUR")
            print(f"  Volatility: {results['period1']['volatility']:.2f}")
            
            print(f"\nPeriod 2 ({results['period2']['start_date']} to {results['period2']['end_date']}, {results['period2']['days']} days):")
            print(f"  Average Rate: {results['period2']['avg_rate']:.2f} HUF/EUR")
            print(f"  Min/Max: {results['period2']['min_rate']:.2f} - {results['period2']['max_rate']:.2f} HUF/EUR")
            print(f"  Volatility: {results['period2']['volatility']:.2f}")
            
            print("\nComparison:")
            print(f"  Average Rate Change: {results['comparison']['avg_rate_diff']:.2f} HUF/EUR ({results['comparison']['avg_rate_pct_change']:.2f}%)")
            print(f"  Volatility Change: {results['comparison']['volatility_change']:.2f} ({results['comparison']['volatility_pct_change']:.2f}%)")
            
    elif args.command == 'sources':
        print("Available data sources:")
        print("  ecb          - European Central Bank")
        print("  mnb          - Hungarian National Bank")
        print("  auto         - Try all sources in priority order")


if __name__ == "__main__":
    main()